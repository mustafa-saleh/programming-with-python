from datetime import datetime
import openpyxl
import requests
from guessing_game import guessing_game
from helper import youngest_employee, num_of_upper_lower_case_letters, even_numbers
from calculator import calculator
from university import Student, Professor, Lecture, Person

employees = [{
    "name": "Tina",
    "age": 30,
    "birthday": "1990-03-10",
    "job": "DevOps Engineer",
    "address": {
        "city": "New York",
        "country": "USA"
    }
},
{
    "name": "Tim",
    "age": 35,
    "birthday": "1985-02-21",
    "job": "Developer",
    "address": {
        "city": "Sydney",
        "country": "Australia"
    }
}]

# Exercise 1 - Lists and Loops
# Write a program that prints out all the elements of the list that are higher than or equal to 10.
# Instead of printing the elements one by one, make a new list that has all the elements higher than or equal to 10 from this list in it and print out this new list.
# Ask the user for a number as input and print a list that contains only those elements from my_list that are higher than the number given by the user.
def exercise_1():
    my_list = [1, 2, 2, 4, 4, 5, 6, 8, 10, 13, 22, 35, 52, 83]
    
    gte10 = [i for i in my_list if i >= 10]
    print(f"greater than or equal to 10, {gte10}")
    
    user_input = input('Enter a number: ')
    gt_num = [i for i in my_list if i > int(user_input)]
    print(gt_num)

# Exercise 2 - Working with Dictionaries
# Write a Python Script that:
# Updates the job to Software Engineer
# Removes the age key from the dictionary
# Loops through the dictionary and prints the key:value pairs one by one
# Merges these two Python dictionaries into 1 new dictionary.
# Sums up all the values in the new dictionary and prints it out
# Prints the max and minimum values of the dictionary
def exercise_2():
    employee = {"name": "Tim", "age": 30, "birthday": "1990-03-10", "job": "DevOps Engineer"}
    dict_one = {'a': 100, 'b': 400}
    dict_two = {'x': 300, 'y': 200}
    
    employee["job"] = "Software Engineer"
    employee.pop("age")
    
    for key, value in employee.items():
        print(f"{key}: {value}")
    
    dict_one_two = dict_one | dict_two
    print(f"merged dict result {dict_one_two}")
    
    sum = 0
    for value in dict_one_two.values():
        sum += value
    
    print(f"sum of values is {sum}")
    
    min_value = min(dict_one_two.values())
    max_value = max(dict_one_two.values())
    print(f"min value is {min_value}, max value is {max_value}")

# Exercise 3 - Dictionaries and Nested Data Structures
# Write a Python Program that:
# Prints out - the name, job and city of each employee using a loop. The program must work for any number of employees in the list, not just 2.
# Prints the country of the second employee in the list by accessing it directly without the loop.
def exercise_3():
    for employee in employees:
        print(f"{employee['name']} works as a {employee['job']} & lives in {employee['address']['city']}")
    
    print(f"country of the 2'nd employee is {employees[1]['address']['country']}")

# Exercise 7 - Object-Oriented Programming
# Imagine you are working in a university and need to write a program, which handles data of students, professors and lectures. 
# To work with this data you create classes and objects:
def exercise_7():
    professor = Professor("x", "y", 60, ["Thermo Dynamics"])
    lecture = Lecture("Thermo Dynamics", 100, 120, [professor])
    student = Student("one", "two", 18, [lecture])
    print(student.full_name())
    print(student.age)
    student.add_lecture(lecture)
    print(student.list_lectures())
    print(professor.full_name())

#Exercise 8 - Date and Time
# Write a program that:
# accepts user's birthday as input
# and calculates how many days, hours and minutes are remaining till the birthday
# prints out the result as a message to the user
def exercise_8():
    today = datetime.now()
    user_birthday = input("Enter your birthday (dd-mm-yyyy): ")
    birthday_date = datetime.strptime(user_birthday, "%d-%m-%Y")
    
    next_birthday = datetime(today.year, birthday_date.month, birthday_date.day)
    if next_birthday < today:
        next_birthday = datetime(today.year + 1, next_birthday.month, next_birthday.day)
    
    delta = next_birthday - today
    print(f"The remaining time till your birthday is {delta.days} days, {delta.seconds // 3600} hours, {delta.seconds % 3600 // 60} minutes")


# Exercise 9 - Working with Spreadsheets
# Write a program that:
# reads the provided spreadsheet file "employees.xlsx" (see Download section at the bottom) with the following information/columns: 
# "name", "years of experience", "job title", "date of birth"
# creates a new spreadsheet file "employees_sorted.xlsx" with following info/columns: "name", "years of experience", 
# where the years of experience is sorted in descending order: so the employee name with the most experience in years is on top.
def exercise_9():
    emp_file = openpyxl.load_workbook('employees.xlsx')
    emp_sheet = emp_file.worksheets[0]
    emp_list = []
    
    for emp in range(2, emp_sheet.max_row + 1):
        name = emp_sheet.cell(row=emp, column=1).value
        years_of_experience = emp_sheet.cell(row=emp, column=2).value
        emp_list.append({"name": name, "years_of_experience": years_of_experience})
    
    sorted_emp_list = sorted(emp_list, key=lambda x: x["years_of_experience"], reverse=True)
    
    for i in range(0, len(sorted_emp_list)):
        emp_sheet.cell(i + 2, 1).value = sorted_emp_list[i]["name"]
        emp_sheet.cell(i + 2, 2).value = sorted_emp_list[i]["years_of_experience"]
        emp_sheet.cell(i + 2, 3).value = ""
        emp_sheet.cell(i + 2, 4).value = ""
    
    emp_file.save('employees_sorted.xlsx')

# Exercise 10 - REST API
# connects to GitHub API
# gets all the public repositories for a specific GitHub user
# prints the name &amp; URL of every project
def exercise_10():
    username = input('Enter your username: ')
    url = f"https://api.github.com/users/{username}/repos"
    response = requests.get(url).json()

    for repo in response:
        print(f"{repo['name']} - {repo['html_url']}")

if __name__ == '__main__':
    # Exercise 10 - REST API
    exercise_10()

    # Exercise 9 - Working with Spreadsheets
    # exercise_9()

    # Exercise 8 - Date and Time
    # exercise_8()

    # Exercise 7 - Object-Oriented Programming
    # exercise_7()

    # Exercise 6 - Guessing Game
    # guessing_game()

    # Exercise 5 - Calculator
    # calculator()

    # Exercise 4 - Helper Functions
    # youngest_employee(employees)
    # num_of_upper_lower_case_letters("hello CAN you hEAr Me")
    # even_numbers([1, 2, 3, 4, 5, 6, 7, 8, 9, 10])

    # Exercise 3 - Dictionaries and Nested Data Structures
    # exercise_3()

    # Exercise 2 - Working with Dictionaries
    # exercise_2()

    # Exercise 1 - Lists and Loops
    # exercise_1()
